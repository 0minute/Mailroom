#-*- coding: utf-8 -*-

from typing import Tuple, List, Dict, Any, Optional
import openpyxl as xl
import pandas as pd
import numpy as np
import os
import sys
import datetime
from dateutil.parser import parse
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
#import win32com.client
#win =win32com.client.Dispatch("Excel.Application")

# Constants
ENCODING = 'CP949'
DATE_FORMAT = '%Y-%m-%d %H:%M:%S'
EXCEL_START_ROW = 4
QUICK_MAX_ROWS = 197
PARCEL_MAX_ROWS = 97

# Column names
ORDER_COLUMNS = [
    '운송장번호', '생성일시', '주문상태', '착불여부', '편도/왕복', '배송수단', 
    '그룹명', '이용사유', '주문자 이름', '출발지 주소', '도착지 주소', '상품', 
    '물품정보(유의사항)', '결제금액', '결제타입'
]

PREMIUM_COLUMNS = ['빌딩명', '이름', '프리미엄 상태', '요금 타입', '계산구조', '계약금액']

# Error messages
ERROR_MESSAGES = {
    'duplicate_company': '중복 입력된 회사명 히스토리(고객사 입력 회사명)가 있습니다. 회사명 테이블을 확인하시어 중복값을 제거해주세요.',
    'unknown_group': '확인되지 않은 고객명이 식별되었습니다. 회사명 테이블을 확인해주세요.',
    'missing_company': '입주사 테이블에 누락된 고객명이 식별되었습니다. 입주사 테이블을 확인해주세요.',
    'missing_building': '소재지 테이블에 누락된 빌딩명이 식별되었습니다. 소재지 테이블을 확인해주세요.'
}

# CU 경로설정
if getattr(sys, 'frozen', False):
    #test.exe로 실행한 경우,test.exe를 보관한 디렉토리의 full path를 취득
    program_directory = os.path.dirname(os.path.abspath(sys.executable))
    cwd = os.path.dirname(os.path.abspath(sys.executable))
    Path_ = os.path.dirname(os.path.abspath(sys.executable))
    Save_path_app = os.path.abspath(os.path.join(Path_, os.pardir))
else:
    #python test.py로 실행한 경우,test.py를 보관한 디렉토리의 full path를 취득
    program_directory = os.path.dirname(os.path.abspath(__file__))
    cwd = os.path.dirname(os.path.abspath(__file__))
    Path_ = os.path.dirname(os.path.abspath(__file__))
    Save_path_app = os.path.dirname(os.path.abspath(__file__))


    
#CU 현재 작업 디렉토리를 변경
os.chdir(program_directory)

#현재 exe또는 py 디렉토리에 저장시키기

## CU 폴더 생성하는 함수
def createDirectory(directory: str) -> None:
    """Create a directory if it doesn't exist.
    
    Args:
        directory: Path of the directory to create.
    """
    try:
        if not os.path.exists(directory):
            os.makedirs(directory)
    except OSError:
        print("Error: Failed to create the directory.")



## CU 테이블 파일이 존재하는지 확인하는 함수
def check_mailTable_exist() -> str:
    """Check if the mail table file exists and return its path.
    
    Returns:
        str: Path to the mail table file or "Error" if not found.
    """
    path_root = program_directory
    print('프로그램실행경로 : ' + path_root)
    valetTablePath = path_root+"/"+'KMPNS_메일룸_Table.xlsx'
    parentPath = os.path.join(path_root, os.pardir)
    valetTableParent = os.path.abspath(parentPath+"/"+'KMPNS_메일룸_Table.xlsx')
    
    if os.path.isfile(valetTablePath) :
        print('테이블 존재를 확인하였습니다.')
        valetTableXl = valetTablePath
        return valetTableXl.replace('\\','/')
    elif os.path.isfile(valetTableParent) :
        print('테이블 존재를 확인하였습니다.')
        valetTableXl = valetTableParent.replace('\\','/')
        return valetTableXl            
    else :
        print('메일룸 테이블이 경로에 없습니다.')
        return "Error"



def check_templateXl_exist():
    path_root = program_directory
    print('프로그램실행경로 : ' + path_root)
    ValettemplPath = path_root+"/"+'KMPNS_메일룸_Template.xlsx'
    parentPath2 = os.path.join(path_root, os.pardir)
    ValettemplParent = os.path.abspath(parentPath2+"/"+'KMPNS_메일룸_Template.xlsx')
        
    if os.path.isfile(ValettemplPath) :
        print('템플릿 존재를 확인하였습니다.')
        ValettemplPath = ValettemplPath.replace('\\','/')
        return ValettemplPath
    elif os.path.isfile(ValettemplParent) :
        print('테이블 존재를 확인하였습니다.')
        ValettemplParent = ValettemplParent.replace('\\','/')
        return ValettemplParent             
    else :
        print('템플릿이 경로에 없습니다.')
        return "Error"
        


# CU Range로 파일 가져오기
class getRngTable:
    """Class to handle Excel range tables.
    
    Attributes:
        file_full_path (str): Full path to the Excel file.
        rngName (str): Name of the range in Excel.
        xls (Workbook): Loaded Excel workbook.
        input_sh_nm (str): Sheet name containing the range.
        cellAddress (str): Cell address of the range.
    """

    def __init__(self, file_full_path: str, rngName: str):
        """Initialize the getRngTable class.
        
        Args:
            file_full_path: Path to the Excel file.
            rngName: Name of the range in Excel.
        """
        self.file_full_path = file_full_path
        self.rngName = rngName
        self.xls = xl.load_workbook(self.file_full_path, read_only=True, keep_vba=True, data_only=True)
        address = list(self.xls.defined_names[rngName].destinations)
        self.input_sh_nm = address[0][0]
        self.cellAddress = address[0][1]

    def input_to_df(self) -> pd.DataFrame:
        """Convert Excel range to pandas DataFrame.
        
        Returns:
            DataFrame containing the data from the Excel range.
        """
        # wb에 input 한 파일 지정
        wb = self.xls
        rngName = self.rngName
        
        # input_sh_nm / cellAddress 변수에 값 지정 및 $ 문자를 address 변수에서 제거
        input_sh_nm = self.input_sh_nm
        cellAddress = self.cellAddress.replace('$','')

        # 엑셀 상 input_data를 list에 추가하기     
        worksheet = wb[input_sh_nm]
        data_rows = []

        for i in range(0,len(worksheet[cellAddress])):
            data_cols = []

            for item in worksheet[cellAddress][i]:
                data_cols.append(item.value)

            # 첫행이 입력된 행까지만 list에 추가
            if data_cols[1] is not None :
                data_rows.append(data_cols)

        # pandas dataframe에 넣어주기
        df = pd.DataFrame(data_rows)

        # Pandas column 설정
        df = df.rename(columns=df.iloc[0]) 
        
        # 첫번째 행 지우기
        df = df.drop(df.index[0]) 

        print(f'{rngName}이 Input되었습니다.')

        return df







# 기본 설정
def get_setting_data(settings_df: pd.DataFrame) -> Tuple[str, str, float, float, int, float]:
    """Get settings data from the settings DataFrame.
    
    Args:
        settings_df: DataFrame containing settings data.
        
    Returns:
        Tuple containing date, report_date, flat_rate, extra_rate, extra_count, and fee.
    """
    date_table = settings_df[settings_df['구분']=='기간']
    report_table = settings_df[settings_df['구분']=='보고일자']
    fix_table = settings_df[settings_df['구분']=='기본요금']
    extra_table = settings_df[settings_df['구분']=='추가요금']
    extra_count_table = settings_df[settings_df['구분']=='추가요금 기준건수']
    fee_table = settings_df[settings_df['구분']=='요금']
    
    date = date_table['값'].iloc[0]
    report_date = report_table['값'].iloc[0]
    flat_rate = fix_table['값'].iloc[0]
    extra_rate = extra_table['값'].iloc[0]
    extra_count = extra_count_table['값'].iloc[0]
    fee = fee_table['값'].iloc[0]
    
    return date, report_date, flat_rate, extra_rate, extra_count, fee


def get_table_mailroom(
    raw_freight_data: pd.DataFrame,
    raw_delivery_data: pd.DataFrame,
    raw_quick_service_data: pd.DataFrame,
    raw_order_data: pd.DataFrame,
    raw_parcel_data: pd.DataFrame,
    table_premium: pd.DataFrame,
    table_name: pd.DataFrame,
    table_account: pd.DataFrame,
    table_price: pd.DataFrame,
    table_formula: pd.DataFrame
) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Process mailroom data and return processed DataFrames."""
    # 헤더 가공
    header_h = raw_freight_data.iloc[0]
    raw_freight_data = raw_freight_data[1:]
    raw_freight_data.columns = header_h

    raw_delivery_rev = raw_delivery_data # 딜리버리 매출 계산용
    
    header_d = raw_delivery_data.iloc[2]
    raw_delivery_data = raw_delivery_data[3:]
    raw_delivery_data.columns = header_d

    # 필요한 열을 선택하여 새로운 DataFrame 생성
    raw_order_data = raw_order_data[ORDER_COLUMNS]

    # 20240716 주문데이터 에서 특정 키워드 제외
    raw_order_data = raw_order_data[raw_order_data['그룹명']!='.']
    raw_order_data = raw_order_data[raw_order_data['착불여부']!='착불']
    raw_order_data = raw_order_data[raw_order_data['주문상태']!='취소']
    raw_order_data = raw_order_data[~raw_order_data['결제타입'].isin(['카드 등록 결제', '현장결제'])]

    # 20240716 주문데이터 전처리
    raw_order_data['그룹명'] = raw_order_data['그룹명'].str.strip()
    raw_order_data.fillna(0,inplace=True)
    
    #20240716 회사명 테이블 정리
    table_name.columns = ['그룹명', '회사명']
    Error_Table_name = table_name[table_name.duplicated(subset = '그룹명')]
    if Error_Table_name['그룹명'].count() != 0:
            with pd.ExcelWriter(Save_path_app + r'/회사명 중복 확인.xlsx') as writer:
                    Error_Table_name.to_excel(writer, sheet_name='회사명중복')
            raise Exception(ERROR_MESSAGES['duplicate_company'])
    
    #20240716 프리미엄 테이블 정리
    table_premium = table_premium[PREMIUM_COLUMNS]
    table_premium = table_premium.fillna('')
    table_premium_on = table_premium[table_premium['프리미엄 상태']=='서비스 중']
    
    #완전성 확인
    raw_order_data = pd.merge(raw_order_data,table_name, on='그룹명', how='left')
    raw_order_data['그룹명'].fillna('Error', inplace=True)
    Error_Table = raw_order_data[raw_order_data['그룹명']=='Error']
    if Error_Table['그룹명'].count() != 0:
            with pd.ExcelWriter(Save_path_app + r'/예외사항_회사명 테이블에 없는 그룹명.xlsx') as writer:
                    Error_Table.to_excel(writer, sheet_name='회사명 테이블 업데이트')
            raise Exception(ERROR_MESSAGES['unknown_group'])
    
    table_premium.columns = ['빌딩명', '회사명', '프리미엄 상태', '요금 타입', '계산구조', '계약금액']
    Error_Compnm = pd.merge(raw_order_data,table_premium, on='회사명', how='left')
    Error_Compnm['빌딩명'].fillna('Error',inplace=True)
    Error_Compnm = Error_Compnm[Error_Compnm['빌딩명']=='Error']
    if Error_Compnm['빌딩명'].count() != 0:
            with pd.ExcelWriter(Save_path_app + r'/예외사항_입주사 테이블에 없는 그룹명.xlsx') as writer:
                    Error_Compnm.to_excel(writer, sheet_name='입주사 테이블 업데이트')
            raise Exception(ERROR_MESSAGES['missing_company'])

    Error_account=pd.merge(table_premium, table_account, on='빌딩명', how='left')
    Error_account['소재지'].fillna('Error',inplace=True)
    Error_account = Error_account[Error_account['소재지']=='Error']
    if Error_account['소재지'].count() != 0:
            with pd.ExcelWriter(Save_path_app + r'/예외사항_소재지 테이블에 없는 빌딩명.xlsx') as writer:
                    Error_Compnm.to_excel(writer, sheet_name='소재지 테이블 업데이트')
            raise Exception(ERROR_MESSAGES['missing_building'])

    # 날짜 데이터 입력

    raw_order_data['생성일시'] = pd.to_datetime(raw_order_data['생성일시'] , format='%Y-%m-%d %H:%M:%S')
    raw_order_data['신청날짜'] = pd.to_datetime(raw_order_data['생성일시']).dt.date
    raw_order_data['신청시간'] = pd.to_datetime(raw_order_data['생성일시']).dt.time

    # 추가필요 열 가공 20240716 필요없어보이지만 남겨둠
    raw_order_data['결제완료여부'] = 0
    raw_order_data['카드, 착불'] = np.where(raw_order_data['결제타입'].str.contains('카드'), 'O', 'X')

            
    # 운송사 데이터 입력
    raw_order_data['운송사'] = np.where(
                            raw_order_data['운송장번호'].str.startswith('3-', na=False), #20240716 2-에서 3-으로 수정
                            '24시화물',
                            np.where(
                                    '택배' == raw_order_data['배송수단'],
                                    '택배', 
                                                        '손자KMC')
                            )
    raw_order_data['마진율'] = np.where(
                            raw_order_data['운송사']=='택배',
                            0.09,
                            np.where(
                                    raw_order_data['운송사']=='손자KMC',
                                    0.23,
                                    np.where(
                                            raw_order_data['운송사']=='24시화물',
                                            0.15,
                                            'Error'
                                            )
                                    )
                            )


    # join 전 가공
    ##택배 Raw 데이터 가공
    # price_dict = {'극소': 4500, '소': 4500, '중': 6000, '대': 9000, '대1': 9000}
    price_dict = table_price.set_index('구분')['금액'].to_dict() # 테이블에서 읽어오도록 함
    raw_parcel_data.rename(columns={'운송장번호':'운송장번호'},inplace=True)
    raw_parcel_data.rename(columns={'박스타입':'운임타입'},inplace=True)
    raw_parcel_data = raw_parcel_data[['운송장번호','운임타입']]
    raw_parcel_data['청구가(부가세제외)_t'] = raw_parcel_data['운임타입'].map(price_dict)


    ## 손자 Raw 데이터 가공
    raw_quick_service_data.rename(columns={'오더번호':'운송장번호'}, inplace=True)
    raw_quick_service_data['청구가(부가세제외)_s'] = raw_quick_service_data['고객적용요금']*(1-raw_quick_service_data['할인율']/100) # 20240730 수정 후 안쓰는 컬럼
    raw_quick_service_data['원가(부가세제외)_s'] = raw_quick_service_data['고객적용요금']*(1-raw_quick_service_data['수수료율']/100)
    raw_quick_service_data = raw_quick_service_data[['운송장번호','청구가(부가세제외)_s','원가(부가세제외)_s','거리-km']]

    ## 화물 Raw 데이터 가공
    raw_freight_data.rename(columns={'화물번호':'운송장번호', '운송료':'원가(부가세제외)_h'}, inplace=True)
    raw_freight_data['청구가(부가세제외)_h'] = np.floor(raw_freight_data['원가(부가세제외)_h'] / 0.85 / 100) * 100 # 20240730 수정 후 안쓰는 컬럼 
    raw_freight_data = raw_freight_data[['운송장번호','원가(부가세제외)_h','차량종류','청구가(부가세제외)_h']]

    # JOIN

    raw_order_data['운송장번호'] = raw_order_data['운송장번호'].astype(str)
    raw_quick_service_data['운송장번호'] = raw_quick_service_data['운송장번호'].astype(str)
    raw_parcel_data['운송장번호'] = raw_parcel_data['운송장번호'].astype(str)

    raw_order_data = pd.merge(raw_order_data,raw_parcel_data, on='운송장번호', how='left')
    raw_order_data = pd.merge(raw_order_data,raw_quick_service_data, on='운송장번호', how='left')
    raw_order_data = pd.merge(raw_order_data,raw_freight_data, on='운송장번호', how='left')
    raw_order_data.fillna(0,inplace=True)

    # JOIN 후 가공

    raw_order_data['운임타입'] = np.where(raw_order_data['배송수단']=='택배',
                            raw_order_data['운임타입'],
                            np.where(                               
                                    raw_order_data['배송수단']=='오토바이 급송',
                                    '급송',
                                    '일반'
                                    )
                            )
    raw_order_data['청구가(부가세제외)'] = np.where(raw_order_data['운송사']=='24시화물', raw_order_data['결제금액']/1.1, # np.where(raw_order_data['운송사']=='24시화물', raw_order_data['청구가(부가세제외)_h'],
                                
                                np.where(raw_order_data['운송사']=='손자KMC', raw_order_data['결제금액']/1.1, #  np.where(raw_order_data['운송사']=='손자KMC', raw_order_data['청구가(부가세제외)_s']
                                np.where(raw_order_data['운송사']=='택배', raw_order_data['청구가(부가세제외)_t'], 'Error')))

    raw_order_data['청구가(부가세제외)'] =raw_order_data['청구가(부가세제외)'].astype(float)

    raw_order_data['원가(부가세제외)'] = np.where(raw_order_data['운송사']=='24시화물', raw_order_data['원가(부가세제외)_h'],
                            np.where(raw_order_data['운송사']=='손자KMC', raw_order_data['원가(부가세제외)_s'],
                            np.where(raw_order_data['운송사']=='택배', raw_order_data['청구가(부가세제외)_t']/1.1, 'Error')))

    raw_order_data['원가(부가세제외)']=raw_order_data['원가(부가세제외)'].astype(float)

    raw_order_data['원가(부가세)'] = np.where(raw_order_data['운송사']=='손자KMC', 0,
                            raw_order_data['원가(부가세제외)']/10)

    raw_order_data['원가(부가세)'] = raw_order_data['원가(부가세)'].astype(float)

    raw_order_data['원가(합계)'] = raw_order_data['원가(부가세제외)'].astype(float) + raw_order_data['원가(부가세)'].astype(float)
    raw_order_data['마진'] = raw_order_data['청구가(부가세제외)'].astype(float) - raw_order_data['원가(부가세)'].astype(float)

    raw_order_data.drop(
            ['청구가(부가세제외)_t',
            '청구가(부가세제외)_s',
            '청구가(부가세제외)_h',
            '원가(부가세제외)_s',
            '원가(부가세제외)_h'],  
            axis=1, inplace=True)

    raw_order_data['차량정보'] = np.where(raw_order_data['운송사']=='택배','택배',
                            np.where(raw_order_data['운송사']=='손자KMC',
                                            np.where(raw_order_data['배송수단'].str.contains('다마스'),'다마스','일반'), #배송수단으로
                            np.where(raw_order_data['운송사']=='24시화물',raw_order_data['차량종류'],'Error')))
    
    Year_ = pd.to_datetime(raw_order_data['신청날짜']).dt.year
    Weak_ = pd.to_datetime(raw_order_data['신청날짜']).dt.strftime('%U')

    raw_order_data['주차'] = Year_.astype(str) + '-W' + Weak_.astype(str)
    raw_order_data['날짜'] = pd.to_datetime(raw_order_data['신청날짜']).dt.day
    raw_order_data_nodata = raw_order_data[raw_order_data['청구가(부가세제외)']==0]
    raw_order_data = raw_order_data[raw_order_data['청구가(부가세제외)']!=0]
    raw_order_data.to_excel(Save_path_app + '\\' + 'RPA_Results.xlsx')  #저장경로
    raw_order_data_nodata.to_excel(Save_path_app + '\\' + 'RPA_Results_NoTransitData.xlsx') #저장경로
    
    
    raw_delivery_report = raw_delivery_data[['년월','년월일','건별','월정액','무료','주차']]
    raw_delivery_report.fillna(0,inplace=True)
    raw_delivery_report = raw_delivery_report[raw_delivery_report['년월']!=0]
    
    # 월정액 매출 계산
    
    header_dm = raw_delivery_rev.iloc[1]
    raw_dm = raw_delivery_rev[2:]
    raw_dm.columns = header_dm
    
    header_d2 = raw_delivery_rev.iloc[2]
    raw_delivery_rev = raw_delivery_rev[3:]
    raw_delivery_rev.columns = header_d2

    raw_d_M=raw_dm[['월정액']]
    raw_d_M.fillna(0,inplace=True)
    
    ## 헤더 추출
    raw_d_M = raw_d_M.iloc[0]
    list_d = raw_d_M.to_list()
    list_d.append('년월')
    
    ## 정액제 계산 대상 회사 추출
    raw_d_fee=raw_delivery_rev[list_d]
    raw_d_fee.fillna(0,inplace=True)
    pivot_d = raw_d_fee.pivot_table(
                                    values=list_d
                                      ,index='년월'
                                      ,aggfunc='sum')
    for M_compn in list_d:
            if M_compn == '년월':
                    pass
            else:
                    pivot_d.loc[pivot_d[M_compn] < extra_count, 'R_' + M_compn] = flat_rate
                    pivot_d.loc[pivot_d[M_compn] >= extra_count, 'R_' + M_compn] = flat_rate + (pivot_d[M_compn]-extra_count)*extra_rate

    ## 합산
    R_list_d = []
    for old_column in list_d:
            if old_column == '년월':
                    pass
            else:
                    R_list_d.append('R_'+old_column)
    pivot_d['월정액_Sum'] = 0
    
    for each_compn in R_list_d:
            pivot_d['월정액_Sum'] =pivot_d['월정액_Sum'] + pivot_d[each_compn]
    
    
    pivot_d.reset_index(inplace=True)

    pivot_d = pivot_d[['년월','월정액_Sum']]
    raw_d_merged = pd.merge(raw_delivery_report,pivot_d,on='년월',how='left')
    raw_d_merged.to_excel(Save_path_app + '\\' + 'RPA_Results_Delivery.xlsx')
    
    return raw_order_data, raw_delivery_data, table_premium, table_premium_on

def process_mailroom(date, report_date, flat_rate, extra_rate, extra_count, fee, processed_order_data, processed_delivery_data, df_premium, df_premium_active, df_formula):    
        # 프리미엄 딜리버리 가공
        processed_delivery_data = processed_delivery_data[processed_delivery_data['년월']==date]
        company_names = processed_order_data['회사명'].unique().tolist()
        company_names_df = pd.DataFrame(processed_order_data['회사명'].unique(), columns = ['이름'])
        company_names_df['포함'] = 'O'
        premium_only = pd.merge(df_premium_active, company_names_df, on='이름', how='left')
        premium_only['포함'].fillna('X',inplace=True)
        premium_only = premium_only[premium_only['포함']=='X']

        # openpyxl
        for company_name in company_names:
                wb_mailroom = xl.load_workbook(dir_template)
                ws_report = wb_mailroom['운영보고서']
                ws_report['C4'] = report_date
                
                # RAW 데이터 붙여넣기                
                filtered_order_data = processed_order_data[processed_order_data['회사명']==company_name]
                filtered_order_data = filtered_order_data[['신청날짜','주문자 이름','이용사유','출발지 주소','도착지 주소','배송수단','편도/왕복','청구가(부가세제외)']]
                
                ws_quick = wb_mailroom['퀵발송']
                quick_delivery_data = filtered_order_data[filtered_order_data['배송수단']!='택배']
                quick_count = quick_delivery_data['배송수단'].count()
                
                if quick_count != 0:
                        header_quick = quick_delivery_data.iloc[0]
                        quick_delivery_data = quick_delivery_data[1:]
                        quick_delivery_data.columns = header_quick
                        
                        for r_idx, row in enumerate(dataframe_to_rows(quick_delivery_data,index=False,header=True), 1):
                                for c_idx, value in enumerate(row, 1):
                                        ws_quick.cell(row=r_idx+3, column=c_idx+1, value=value)
                
                target_row = 4
                end_row_quick = quick_count + target_row
                ws_quick.delete_rows(end_row_quick, 197-quick_count - 1)        
                ws_quick['I'+str(quick_count+7-1)] = '=SUM(I4:I' + str(quick_count+6-2) +')'
                ws_quick['I'+str(quick_count+7)] = '=' + 'I' + str(quick_count+7+1-2) + '*10%'
                ws_quick['I'+str(quick_count+7-2)] = '=SUM(' + 'I'+str(quick_count+7+1-2) + ':' + 'I'+str(quick_count+7+2-2) + ')'
                                
                ws_parcel = wb_mailroom['택배발송']
                parcel_delivery_data = filtered_order_data[filtered_order_data['배송수단']=='택배']
                parcel_count = parcel_delivery_data['배송수단'].count()
                
                if parcel_count != 0:
                        header_parcel = parcel_delivery_data.iloc[0]
                        parcel_delivery_data = parcel_delivery_data[1:]
                        parcel_delivery_data.columns = header_parcel
                
                        for r_idx, row in enumerate(dataframe_to_rows(parcel_delivery_data,index=False,header=True), 1):
                                for c_idx, value in enumerate(row, 1):
                                        ws_parcel.cell(row=r_idx+3, column=c_idx+1, value=value)                   
                
                target_row = 4
                end_row_parcel = parcel_count + target_row
                ws_parcel.delete_rows(end_row_parcel, 97-parcel_count - 1)  
                ws_parcel['I'+str(parcel_count+7+1-2)] = '=SUM(I4:I' + str(parcel_count+6-2) +')'
                ws_parcel['I'+str(parcel_count+7+2-2)] = '=' + 'I' + str(parcel_count+7+1-2) + '*10%'
                ws_parcel['I'+str(parcel_count+7-2)] = '=SUM(' + 'I'+str(parcel_count+7+1-2) + ':' + 'I'+str(parcel_count+7+2-2) + ')'
                
                ws_report['F12'] = '=퀵발송!I' + str(quick_count+7-1)
                ws_report['F13'] = '=택배발송!I' + str(parcel_count+7-1)
                premium_tower = df_premium[df_premium['회사명']==company_name]
                tower_name = premium_tower['빌딩명'].iloc[0]
                
                ws_report['B2']= tower_name +' 얼른 딜리버리 사용 내역서'
                ws_report['C5']= df_account[df_account['빌딩명']==tower_name]['소재지'].iloc[0]
                
                ws_delivery = wb_mailroom['일자별배송건수']
                ws_delivery['C3']=company_name
                
                filtered_premium = df_premium_active[df_premium_active['이름']==company_name]
                if filtered_premium['이름'].count() != 0:
                        # ... rest of the code ...


if __name__ == '__main__':
        
        #Input파일
        dir_freight_data = r'\input\Data_24시화물.xlsx'
        dir_delivery_data = r'\input\Data_딜리버리.xlsx'
        dir_quick_service_data = r'\input\Data_손자.csv'
        dir_order_data = r'\input\Data_신청내역.xlsx'
        dir_parcel_data = r'\input\Data_택배2.xlsx'
        
        #기본파일
        mail_table_xl = 'KMPNS_메일룸_Table.xlsx'
        dir_template = 'KMPNS_메일룸_Template.xlsx'
        
        table_premium = getRngTable(mail_table_xl, 'Company')    #'\Table_입주사.xlsx'
        table_name = getRngTable(mail_table_xl, 'GoogleName')    #'\Table_회사명.xlsx'
        table_account = getRngTable(mail_table_xl, 'Location')   #'\Table_소재지.csv'
        table_setting = getRngTable(mail_table_xl, 'Setting')    #'\Table_설정.xlsx'
        table_price = getRngTable(mail_table_xl, 'Price')    #20240716 규격별 택배비 테이블
        table_formula = getRngTable(mail_table_xl, 'Formula')

        raw_freight_data = pd.read_excel(Path_+dir_freight_data)
        raw_delivery_data = pd.read_excel(Path_+dir_delivery_data)
        raw_quick_service_data = pd.read_csv(Path_+dir_quick_service_data, encoding='CP949')
        raw_order_data = pd.read_excel(Path_+dir_order_data)  
        raw_parcel_data = pd.read_excel(Path_+dir_parcel_data)
        df_premium = table_premium.input_to_df()    
        df_name = table_name.input_to_df()          
        df_account = table_account.input_to_df()    
        df_settings = table_setting.input_to_df()        
        df_price = table_price.input_to_df()
        df_formula = table_formula.input_to_df()

        createDirectory(Path_+'\\'+datetime.datetime.now().strftime("%Y%m%d_%H%M%S"))
                
        save_path = Path_+'\\'+datetime.datetime.now().strftime("%Y%m%d_%H%M%S")

        date, report_date, flat_rate, extra_rate, extra_count, fee = get_setting_data(df_settings)
        processed_order_data, processed_delivery_data, df_premium, df_premium_active = get_table_mailroom(raw_freight_data, raw_delivery_data, raw_quick_service_data, raw_order_data, raw_parcel_data, df_premium, df_name, df_account, df_price, df_formula)
        process_mailroom(date, report_date, flat_rate, extra_rate, extra_count, fee, processed_order_data, processed_delivery_data, df_premium, df_premium_active, df_formula)
        
        
        